from __future__ import annotations

from datetime import datetime, timezone
import json
import os
from pathlib import Path
from threading import Lock
from typing import Any

from openpyxl import Workbook, load_workbook


REPORT_DIR = Path("reports")
REPORT_FILE = REPORT_DIR / "issue_reports.xlsx"
DEFAULT_GOOGLE_SHEET_ID = "1Mar6ZRn1rOG8AH1O_CGMVs5v5xrLDqdaEpbhroX0Xvw"
REPORT_HEADERS = [
    "Timestamp UTC",
    "Molecule Query",
    "Report Type",
    "Message",
    "Contact",
    "Page URL",
    "User Agent",
]
GOOGLE_SCOPES = ["https://www.googleapis.com/auth/spreadsheets"]
GOOGLE_WORKSHEET_NAME = os.getenv("GOOGLE_SHEET_WORKSHEET", "Issue Reports")

_report_lock = Lock()


def save_issue_report(report: dict[str, Any]) -> dict[str, str]:
    row = _build_report_row(report)
    google_sheet_id = os.getenv("GOOGLE_SHEET_ID", DEFAULT_GOOGLE_SHEET_ID).strip()

    if google_sheet_id:
        _save_issue_report_to_google_sheet(google_sheet_id, row)
        return {"destination": "google_sheet", "saved_to": google_sheet_id}

    saved_path = _save_issue_report_to_excel(row)
    return {"destination": "excel", "saved_to": str(saved_path)}


def _build_report_row(report: dict[str, Any]) -> list[str]:
    return [
        datetime.now(timezone.utc).isoformat(timespec="seconds"),
        str(report.get("molecule_query", "")),
        str(report.get("report_type", "")),
        str(report.get("message", "")),
        str(report.get("contact", "")),
        str(report.get("page_url", "")),
        str(report.get("user_agent", "")),
    ]


def _save_issue_report_to_google_sheet(sheet_id: str, row: list[str]) -> None:
    import gspread

    credentials = _get_google_credentials()
    client = gspread.authorize(credentials)
    spreadsheet = client.open_by_key(sheet_id)

    try:
        sheet = spreadsheet.worksheet(GOOGLE_WORKSHEET_NAME)
    except gspread.WorksheetNotFound:
        sheet = spreadsheet.add_worksheet(
            title=GOOGLE_WORKSHEET_NAME,
            rows=1000,
            cols=len(REPORT_HEADERS),
        )

    existing_headers = sheet.row_values(1)
    if existing_headers != REPORT_HEADERS:
        sheet.clear()
        sheet.append_row(REPORT_HEADERS, value_input_option="RAW")

    sheet.append_row(row, value_input_option="RAW")


def _get_google_credentials() -> Credentials:
    from google.oauth2.service_account import Credentials

    raw_json = os.getenv("GOOGLE_SERVICE_ACCOUNT_JSON", "").strip()
    credentials_path = os.getenv("GOOGLE_APPLICATION_CREDENTIALS", "").strip()

    if raw_json:
        service_account_info = json.loads(raw_json)
        return Credentials.from_service_account_info(
            service_account_info,
            scopes=GOOGLE_SCOPES,
        )

    if credentials_path:
        return Credentials.from_service_account_file(
            credentials_path,
            scopes=GOOGLE_SCOPES,
        )

    raise RuntimeError("Google Sheets credentials are not configured.")


def _save_issue_report_to_excel(row: list[str]) -> Path:
    REPORT_DIR.mkdir(parents=True, exist_ok=True)

    with _report_lock:
        if REPORT_FILE.exists():
            workbook = load_workbook(REPORT_FILE)
            sheet = workbook.active
            if sheet.max_row == 0:
                sheet.append(REPORT_HEADERS)
        else:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Issue Reports"
            sheet.append(REPORT_HEADERS)

        sheet.append(row)

        for column_cells in sheet.columns:
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = min(
                max(max_length + 2, 14),
                60,
            )

        workbook.save(REPORT_FILE)

    return REPORT_FILE
